"""Fill Wayfair templates for adding new wallpaper variants to existing listings."""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import json
import os
from pathlib import Path
import re
import time
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from data.pricing import PriceProvider
from data.wallpaper_shaper import WallpaperDataShaper

# python add_wallpaper_variants.py "/Users/danielkravchenko/Downloads/Product Addition Template - [05_29_2026].xlsx" \
#     --dropbox-folder "/Wayfair Wallpaper" \
#     -o /Users/danielkravchenko/Downloads/wayfair_variants_dropbox_test.xlsx

DATA_START_ROW = 7
HEADER_ROW = 4
REQUIRED_NEW_MATERIALS = ("Peel-n-Stick: Canvas", "Non-Woven: Premium")
DEFAULT_MULTICOLOR_VALUE = "Multicolor"
MATERIAL_GROUPING_LABELS = {
    "Peel-n-Stick": "Peel & Stick",
    "Non-Woven": "Non-woven",
    "Peel-n-Stick: Canvas": "Luxury Canvas Texture",
    "Non-Woven: Premium": "Premium Non-Woven",
}
DROPBOX_API_BASE_URL = "https://api.dropboxapi.com/2"
DEFAULT_DROPBOX_FOLDER = "/Wayfair Wallpaper"
DROPBOX_TOKEN_ENV_VAR = "DROPBOX_TOKEN"
DROPBOX_TOKEN = ""
DROPBOX_APP_KEY_ENV_VAR = "DROPBOX_APP_KEY"
DROPBOX_APP_SECRET_ENV_VAR = "DROPBOX_APP_SECRET"
DROPBOX_REFRESH_TOKEN_ENV_VAR = "DROPBOX_REFRESH_TOKEN"
DROPBOX_APP_KEY = ""
DROPBOX_APP_SECRET = ""
DROPBOX_REFRESH_TOKEN = ""
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff")
LOW_PRIORITY_IMAGE_NAME_TOKENS = ("copy", "duplicate", "old", "backup")
REQUEST_RETRY_COUNT = 3
REQUEST_RETRY_DELAY_SECONDS = 2
RETRYABLE_HTTP_STATUSES = {429, 500, 502, 503, 504}
SERVICE_SHEET_NAMES = {
    "Instructions",
    "Error Summary",
    "Valid Values",
    "WF_EXISTING_PARTS",
    "WAYFAIR_USE_ONLY",
    "Additional Cartons",
    "Additional Chemicals",
    "Additional Images",
    "Additional Videos",
    "Additional Documents",
}
PART_NUMBER_PATTERN = re.compile(
    r"^(?P<base>.+?)\s+(?P<width>\d+(?:\.\d+)?)\s*x\s*(?P<height>\d+(?:\.\d+)?)\s+(?P<suffix>.+)$",
    re.IGNORECASE,
)
SIZE_LABEL_PATTERN = re.compile(
    r'(?P<length>\d+(?:\.\d+)?)"\s*L\s*x\s*(?P<width>\d+(?:\.\d+)?)"\s*W',
    re.IGNORECASE,
)


@dataclass(frozen=True)
class ExistingVariant:
    """Existing variant data parsed from one SKU sheet row."""

    row_number: int
    part_number: str
    base_code: str
    width: int
    height: int
    material_name: str | None


@dataclass(frozen=True)
class SheetContext:
    """Reusable data inferred from one SKU sheet."""

    sheet_name: str
    headers: dict[str, int]
    existing_variants: list[ExistingVariant]
    product_name: str
    base_code: str
    image_links: list[str]
    color: str
    commercial_warranty: str | None
    commercial_warranty_length: str | None
    dropbox_image_missing: bool = False
    dropbox_image_error: str | None = None


@dataclass(frozen=True)
class FillWarning:
    """Non-fatal issue found while filling a template."""

    sheet_name: str
    message: str


@dataclass(frozen=True)
class DropboxFile:
    """Dropbox file metadata required for image matching and sharing."""

    name: str
    path: str


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser."""

    parser = argparse.ArgumentParser(
        description=(
            "Fill Wayfair New Variants templates with additional wallpaper "
            "material variants."
        )
    )
    parser.add_argument("template", type=Path, help="Path to the Wayfair .xlsx file.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output .xlsx path. Defaults to '<SKU>_<SKU>_variants.xlsx' near the input file.",
    )
    parser.add_argument(
        "--materials",
        nargs="+",
        default=list(REQUIRED_NEW_MATERIALS),
        help="Wallpaper material variants to add.",
    )
    parser.add_argument(
        "--image-map",
        type=Path,
        help=(
            "Optional CSV with sheet/base_sku/sku and image1..image5 columns. "
            "Images from the CSV override images found in the workbook."
        ),
    )
    parser.add_argument(
        "--dropbox-folder",
        default=DEFAULT_DROPBOX_FOLDER,
        help=(
            "Dropbox folder path with main images. The script finds the first "
            "image whose file name starts with the base SKU, for example VN006."
        ),
    )
    parser.add_argument(
        "--dropbox-token",
        help=(
            "Short-lived Dropbox API access token. Defaults to "
            f"${DROPBOX_TOKEN_ENV_VAR}."
        ),
    )
    parser.add_argument(
        "--dropbox-app-key",
        help=(
            "Dropbox app key for refresh-token auth. Defaults to "
            f"${DROPBOX_APP_KEY_ENV_VAR}."
        ),
    )
    parser.add_argument(
        "--dropbox-app-secret",
        help=(
            "Dropbox app secret for refresh-token auth. Defaults to "
            f"${DROPBOX_APP_SECRET_ENV_VAR}."
        ),
    )
    parser.add_argument(
        "--dropbox-refresh-token",
        help=(
            "Dropbox refresh token. Defaults to "
            f"${DROPBOX_REFRESH_TOKEN_ENV_VAR}."
        ),
    )
    parser.add_argument(
        "--dropbox-no-recursive",
        action="store_true",
        help="Do not search Dropbox subfolders. Shared folder URLs are always non-recursive.",
    )
    parser.add_argument(
        "--dropbox-list-files",
        action="store_true",
        help="List image files visible through --dropbox-folder and exit.",
    )
    parser.add_argument(
        "--price-csv-url",
        default=None,
        help="Optional Google Sheets CSV export URL for prices.",
    )
    parser.add_argument(
        "--default-color",
        default=DEFAULT_MULTICOLOR_VALUE,
        help="Color value used when the template has no existing Color answer.",
    )
    parser.add_argument(
        "--fill-existing",
        action="store_true",
        help="Also fill required fields for already existing part-number rows.",
    )
    parser.add_argument(
        "--write-additional-images",
        action="store_true",
        help=(
            "Write overflow technical images to the Additional Images sheet. "
            "Disabled by default to avoid exceeding the template row limit."
        ),
    )
    parser.add_argument(
        "--skip-technical-images",
        action="store_true",
        default=True,
        help="Do not write default wallpaper technical images to image slots.",
    )
    parser.add_argument(
        "--include-technical-images",
        action="store_false",
        dest="skip_technical_images",
        help="Write default wallpaper technical images to image slots.",
    )
    return parser


def load_image_map(path: Path | None) -> dict[str, list[str]]:
    """Load optional per-listing image links from a CSV file."""

    if path is None:
        return {}

    image_map: dict[str, list[str]] = {}
    with path.open(newline="", encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            keys = [
                row.get("sheet", ""),
                row.get("sku", ""),
                row.get("base_sku", ""),
                row.get("base_code", ""),
            ]
            links = [
                row.get(f"image{index}", "") or row.get(f"image_{index}", "")
                for index in range(1, 6)
            ]
            cleaned_links = [link.strip() for link in links if link and link.strip()]
            for key in keys:
                normalized_key = key.strip()
                if normalized_key:
                    image_map[normalized_key] = cleaned_links

    return image_map


def normalize_lookup_key(value: str) -> str:
    """Normalize a SKU or file name for prefix matching."""

    return re.sub(r"[^a-z0-9]", "", value.lower())


def make_lookup_variants(value: str) -> set[str]:
    """Return order-tolerant lookup keys for SKU/file-name matching."""

    normalized = normalize_lookup_key(value)
    if not normalized:
        return set()

    variants = {normalized}
    letters = "".join(re.findall(r"[a-z]+", normalized))
    digits = "".join(re.findall(r"\d+", normalized))
    if letters and digits:
        variants.add(f"{letters}{digits}")
        variants.add(f"{digits}{letters}")
    return variants


def normalize_dropbox_url(url: str) -> str:
    """Normalize Dropbox shared links to direct file URLs Wayfair can download."""

    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.pop("dl", None)
    query["raw"] = "1"
    return urlunsplit(
        (parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment)
    )


def normalize_image_url(url: str) -> str:
    """Return a downloadable URL for supported image hosts."""

    parts = urlsplit(url)
    if parts.netloc.lower().endswith("dropbox.com"):
        return normalize_dropbox_url(url)
    return url


def normalize_image_urls(urls: list[str]) -> list[str]:
    """Normalize image URLs before writing them to Wayfair cells."""

    return [normalize_image_url(url) for url in urls]


def read_http_request(request: Request, description: str) -> bytes:
    """Run an HTTP request with short retries for transient network failures."""

    last_error: Exception | None = None
    for attempt in range(1, REQUEST_RETRY_COUNT + 1):
        try:
            with urlopen(request, timeout=30) as response:
                return response.read()
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            if (
                exc.code not in RETRYABLE_HTTP_STATUSES
                or attempt == REQUEST_RETRY_COUNT
            ):
                raise RuntimeError(f"{description}: HTTP {exc.code}: {detail}") from exc
            last_error = exc
        except (OSError, URLError) as exc:
            if attempt == REQUEST_RETRY_COUNT:
                raise RuntimeError(f"{description}: {exc}") from exc
            last_error = exc
        time.sleep(REQUEST_RETRY_DELAY_SECONDS)

    raise RuntimeError(f"{description}: {last_error}")


def score_image_match(file: DropboxFile, lookup_prefixes: set[str]) -> tuple[int, str]:
    """Score a Dropbox image candidate for SKU matching."""

    stem = Path(file.name).stem
    stem_key = normalize_lookup_key(stem)
    file_keys = make_lookup_variants(stem)
    matching_suffixes = [
        file_key.removeprefix(prefix)
        for prefix in lookup_prefixes
        for file_key in file_keys
        if file_key.startswith(prefix)
    ]
    if not matching_suffixes:
        return (10_000, file.path.lower())

    suffix = min(matching_suffixes, key=len)
    score = len(suffix)
    if any(token in stem_key for token in LOW_PRIORITY_IMAGE_NAME_TOKENS):
        score += 100
    if suffix and not re.fullmatch(r"[a-z]?\d?", suffix):
        score += 20
    return (score, file.path.lower())


class DropboxImageResolver:
    """Resolve main image links from a Dropbox folder via Dropbox HTTP API."""

    def __init__(
        self,
        token: str,
        folder_path: str,
        recursive: bool = True,
    ) -> None:
        """Store Dropbox API settings."""

        self.token = token
        self.folder_path = folder_path
        self.recursive = recursive
        self.shared_folder_url = folder_path if folder_path.startswith("http") else None
        self.files: list[DropboxFile] | None = None
        self.shared_links: dict[str, str] = {}
        self.last_match_error: str | None = None

    def resolve_main_image(self, *prefixes: str) -> str | None:
        """Return a shared image link for the first file matching a SKU prefix."""

        self.last_match_error = None
        files = self.list_image_files()
        lookup_prefixes = {
            lookup_variant
            for prefix in prefixes
            for lookup_variant in make_lookup_variants(prefix)
        }
        candidates = sorted(
            (
                file
                for file in files
                if score_image_match(file, lookup_prefixes)[0] < 10_000
            ),
            key=lambda file: score_image_match(file, lookup_prefixes),
        )
        for file in candidates:
            try:
                return self.get_or_create_shared_link(file.path)
            except RuntimeError as exc:
                self.last_match_error = f"{file.name}: {exc}"
                continue
        return None

    def list_image_files(self) -> list[DropboxFile]:
        """Return cached image files found in the configured Dropbox folder."""

        if self.files is not None:
            return self.files

        entries: list[DropboxFile] = []
        response = self.api_call("files/list_folder", self.make_list_folder_payload())
        entries.extend(self.parse_file_entries(response))
        while response.get("has_more"):
            response = self.api_call(
                "files/list_folder/continue",
                {"cursor": response["cursor"]},
            )
            entries.extend(self.parse_file_entries(response))

        self.files = sorted(entries, key=lambda file: file.path.lower())
        return self.files

    def print_visible_files(self, limit: int = 200) -> None:
        """Print image files visible through the configured Dropbox folder."""

        files = self.list_image_files()
        print(f"Dropbox image files found: {len(files)}")
        for file in files[:limit]:
            print(f"- {file.name} | {file.path}")
        if len(files) > limit:
            print(f"... {len(files) - limit} more")

    def make_list_folder_payload(self) -> dict[str, Any]:
        """Build the Dropbox list_folder payload for a path or shared folder URL."""

        payload: dict[str, Any] = {
            "recursive": False if self.shared_folder_url else self.recursive,
            "include_non_downloadable_files": False,
        }
        if self.shared_folder_url:
            payload["path"] = ""
            payload["shared_link"] = {"url": self.shared_folder_url}
        else:
            payload["path"] = self.folder_path
        return payload

    def get_or_create_shared_link(self, path: str) -> str:
        """Return an existing shared link or create a new one for a file."""

        if path in self.shared_links:
            return self.shared_links[path]

        if self.shared_folder_url:
            url = self.get_nested_shared_link(path)
            self.shared_links[path] = url
            return url

        existing_link = self.get_existing_shared_link(path)
        if existing_link:
            self.shared_links[path] = existing_link
            return existing_link

        try:
            response = self.api_call(
                "sharing/create_shared_link_with_settings",
                {"path": path, "settings": {"requested_visibility": "public"}},
            )
            url = normalize_dropbox_url(response["url"])
            self.shared_links[path] = url
            return url
        except RuntimeError as exc:
            if "shared_link_already_exists" not in str(exc):
                raise
            existing_link = self.get_existing_shared_link(path)
            if existing_link:
                self.shared_links[path] = existing_link
                return existing_link
            raise

    def get_nested_shared_link(self, path: str) -> str:
        """Return a shared link for a file inside a shared folder URL."""

        if not self.shared_folder_url:
            raise RuntimeError("Dropbox shared folder URL is not configured.")
        response = self.api_call(
            "sharing/get_shared_link_metadata",
            {"url": self.shared_folder_url, "path": path},
        )
        url = response.get("url")
        if not isinstance(url, str):
            raise RuntimeError(f"Dropbox did not return a shared link for {path}.")
        return normalize_dropbox_url(url)

    def get_existing_shared_link(self, path: str) -> str | None:
        """Return the first direct shared link for a Dropbox file path."""

        response = self.api_call(
            "sharing/list_shared_links",
            {"path": path, "direct_only": True},
        )
        links = response.get("links", [])
        if not links:
            return None
        url = links[0].get("url")
        return normalize_dropbox_url(url) if isinstance(url, str) else None

    def api_call(self, endpoint: str, payload: dict[str, Any]) -> dict[str, Any]:
        """Call one Dropbox JSON endpoint and return its decoded response."""

        request = Request(
            url=f"{DROPBOX_API_BASE_URL}/{endpoint}",
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {self.token}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        response_body = read_http_request(
            request,
            f"Dropbox API request failed: {endpoint}",
        )
        return json.loads(response_body.decode("utf-8"))

    @staticmethod
    def parse_file_entries(response: dict[str, Any]) -> list[DropboxFile]:
        """Extract image file entries from a Dropbox list_folder response."""

        files: list[DropboxFile] = []
        for entry in response.get("entries", []):
            if entry.get(".tag") != "file":
                continue
            name = entry.get("name")
            path = entry.get("path_lower") or entry.get("path_display")
            if not isinstance(name, str) or not isinstance(path, str):
                continue
            if name.lower().endswith(IMAGE_EXTENSIONS):
                files.append(DropboxFile(name=name, path=path))
        return files


def get_refreshed_dropbox_token(
    app_key: str | None,
    app_secret: str | None,
    refresh_token: str | None,
) -> str | None:
    """Return a fresh Dropbox access token using a stored refresh token."""

    if not app_key or not app_secret or not refresh_token:
        return None

    payload = urlencode(
        {
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "client_id": app_key,
            "client_secret": app_secret,
        }
    ).encode("utf-8")
    request = Request(
        url="https://api.dropboxapi.com/oauth2/token",
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    response_body = read_http_request(request, "Dropbox refresh token request failed")
    data = json.loads(response_body.decode("utf-8"))

    access_token = data.get("access_token")
    if not isinstance(access_token, str) or not access_token:
        raise RuntimeError("Dropbox did not return a refreshed access token.")
    return access_token


def make_dropbox_resolver(
    folder_path: str | None,
    token: str | None,
    app_key: str | None,
    app_secret: str | None,
    refresh_token: str | None,
    recursive: bool,
) -> DropboxImageResolver | None:
    """Create a Dropbox resolver when the command line enables it."""

    if not folder_path:
        return None
    resolved_token = token or DROPBOX_TOKEN or os.getenv(DROPBOX_TOKEN_ENV_VAR)
    if not resolved_token:
        resolved_token = get_refreshed_dropbox_token(
            app_key=app_key or DROPBOX_APP_KEY or os.getenv(DROPBOX_APP_KEY_ENV_VAR),
            app_secret=app_secret
            or DROPBOX_APP_SECRET
            or os.getenv(DROPBOX_APP_SECRET_ENV_VAR),
            refresh_token=refresh_token
            or DROPBOX_REFRESH_TOKEN
            or os.getenv(DROPBOX_REFRESH_TOKEN_ENV_VAR),
        )
    if not resolved_token:
        raise ValueError(
            "Dropbox folder was provided, but no token was found. "
            "Fill DROPBOX_REFRESH_TOKEN, DROPBOX_APP_KEY, and DROPBOX_APP_SECRET "
            "in code, pass them as CLI flags, or set matching environment variables."
        )
    return DropboxImageResolver(
        token=resolved_token,
        folder_path=folder_path,
        recursive=recursive,
    )


def safe_filename_part(value: str) -> str:
    """Return a filesystem-safe filename segment."""

    return re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_")


def collect_sku_sheet_names(input_path: Path) -> list[str]:
    """Return SKU sheet names from a Wayfair workbook."""

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        return [
            worksheet.title
            for worksheet in workbook.worksheets
            if is_sku_sheet(worksheet)
        ]
    finally:
        workbook.close()


def default_output_path(input_path: Path) -> Path:
    """Return the default output path for a filled workbook."""

    sku_names = [
        safe_filename_part(name) for name in collect_sku_sheet_names(input_path)
    ]
    if sku_names:
        return input_path.with_name(
            f"{'_'.join(sku_names)}_variants{input_path.suffix}"
        )
    return input_path.with_name(f"{input_path.stem}_variants{input_path.suffix}")


def make_headers(ws: Worksheet) -> dict[str, int]:
    """Map header names to one-based column indexes."""

    headers: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(HEADER_ROW, column).value
        if isinstance(value, str) and value.strip():
            headers[value.strip()] = column
    return headers


def get_cell_value(
    ws: Worksheet, headers: dict[str, int], row: int, header: str
) -> Any:
    """Read a cell by row and header name."""

    column = headers.get(header)
    if column is None:
        return None
    return ws.cell(row, column).value


def set_cell_value(
    ws: Worksheet,
    headers: dict[str, int],
    row: int,
    header: str,
    value: Any,
) -> None:
    """Write a cell when the target header exists."""

    column = headers.get(header)
    if column is not None:
        ws.cell(row, column).value = value


def is_sku_sheet(ws: Worksheet) -> bool:
    """Return whether a worksheet is a SKU data sheet."""

    if ws.title in SERVICE_SHEET_NAMES or "-WUO_" in ws.title:
        return False
    headers = make_headers(ws)
    return "Supplier Part Number" in headers and "Product Name" in headers


def parse_part_number(part_number: str) -> tuple[str, int, int] | None:
    """Parse base code and dimensions from a supplier part number."""

    match = PART_NUMBER_PATTERN.match(part_number.strip())
    if not match:
        return None
    return (
        re.sub(r"\s+", " ", match.group("base").strip()),
        int(float(match.group("width"))),
        int(float(match.group("height"))),
    )


def parse_size_label(size_label: str) -> tuple[int, int] | None:
    """Parse Wayfair size labels formatted as '<height>\" L x <width>\" W'."""

    match = SIZE_LABEL_PATTERN.search(size_label.strip())
    if not match:
        return None
    return int(float(match.group("width"))), int(float(match.group("length")))


def normalize_material_name(value: str) -> str | None:
    """Map Wayfair display material names to internal material keys."""

    normalized = value.strip().lower().replace("&", "and")
    if "canvas" in normalized:
        return "Peel-n-Stick: Canvas"
    if "premium" in normalized:
        return "Non-Woven: Premium"
    if "peel" in normalized and "stick" in normalized:
        return "Peel-n-Stick"
    if "non" in normalized and "woven" in normalized:
        return "Non-Woven"
    return None


def material_grouping_label(
    material_name: str,
    shaper: WallpaperDataShaper,
) -> str:
    """Return a stable variant label for a wallpaper material."""

    return MATERIAL_GROUPING_LABELS.get(
        material_name,
        shaper.print_type[material_name].display_name,
    )


def find_existing_variants(
    ws: Worksheet,
    headers: dict[str, int],
) -> list[ExistingVariant]:
    """Parse existing part-number rows from a SKU sheet."""

    variants: list[ExistingVariant] = []
    part_column = headers["Supplier Part Number"]
    for row in range(DATA_START_ROW, ws.max_row + 1):
        part_number = ws.cell(row, part_column).value
        if not isinstance(part_number, str) or not part_number.strip():
            continue

        parsed_part = parse_part_number(part_number)
        size_value = get_variant_value(ws, headers, row, "Size")
        parsed_size = (
            parse_size_label(size_value) if isinstance(size_value, str) else None
        )
        if parsed_part is None and parsed_size is None:
            continue

        material_name = normalize_material_name(part_number)
        material_value = get_variant_value(ws, headers, row, "Wallpaper Material")
        if material_name is None:
            material_name = (
                normalize_material_name(material_value)
                if isinstance(material_value, str)
                else None
            )

        if parsed_part is not None:
            base_code, width, height = parsed_part
        else:
            base_code = part_number.split(maxsplit=1)[0]
            assert parsed_size is not None
            width, height = parsed_size

        variants.append(
            ExistingVariant(
                row_number=row,
                part_number=part_number.strip(),
                base_code=base_code,
                width=width,
                height=height,
                material_name=material_name,
            )
        )

    return variants


def get_variant_value(
    ws: Worksheet,
    headers: dict[str, int],
    row: int,
    grouping_name: str,
) -> str | None:
    """Return the value for a named variant grouping in a row."""

    for index in range(1, 4):
        grouping = get_cell_value(ws, headers, row, f"Variant Grouping {index}")
        if is_variant_grouping(grouping, grouping_name):
            value = get_cell_value(
                ws,
                headers,
                row,
                f"Variant Attribute Name On Site {index}",
            )
            return value if isinstance(value, str) else None
    return None


def is_variant_grouping(value: Any, expected: str) -> bool:
    """Return whether a cell value matches a variant grouping name."""

    if not isinstance(value, str):
        return False
    return normalize_lookup_key(value) == normalize_lookup_key(expected)


def collect_existing_images(
    ws: Worksheet,
    headers: dict[str, int],
    existing_variants: list[ExistingVariant],
) -> list[str]:
    """Collect image links from the first existing row that has media values."""

    for variant in existing_variants:
        links = [
            get_cell_value(
                ws, headers, variant.row_number, f"Image File Name or URL {index}"
            )
            for index in range(1, 6)
        ]
        cleaned_links = [
            link.strip() for link in links if isinstance(link, str) and link.strip()
        ]
        if cleaned_links:
            return cleaned_links
    return []


def first_non_empty_value(
    ws: Worksheet,
    headers: dict[str, int],
    existing_variants: list[ExistingVariant],
    header: str,
) -> str | None:
    """Return the first non-empty string value from existing rows."""

    for variant in existing_variants:
        value = get_cell_value(ws, headers, variant.row_number, header)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def make_sheet_context(
    ws: Worksheet,
    image_map: dict[str, list[str]],
    dropbox_resolver: DropboxImageResolver | None,
    default_color: str,
) -> SheetContext | None:
    """Build context needed to generate new rows for one SKU sheet."""

    headers = make_headers(ws)
    existing_variants = find_existing_variants(ws, headers)
    if not existing_variants:
        return None

    product_name = first_non_empty_value(
        ws,
        headers,
        existing_variants,
        "Product Name",
    )
    if not product_name:
        return None

    base_code = existing_variants[0].base_code
    dropbox_image = (
        dropbox_resolver.resolve_main_image(base_code, ws.title)
        if dropbox_resolver is not None
        else None
    )
    dropbox_image_missing = dropbox_resolver is not None and dropbox_image is None
    dropbox_image_error = (
        dropbox_resolver.last_match_error if dropbox_resolver is not None else None
    )
    image_links = (
        image_map.get(ws.title)
        or image_map.get(base_code)
        or ([dropbox_image] if dropbox_image else [])
        or collect_existing_images(ws, headers, existing_variants)
    )
    color = (
        first_non_empty_value(ws, headers, existing_variants, "Color") or default_color
    )
    commercial_warranty = first_non_empty_value(
        ws, headers, existing_variants, "Commercial Warranty"
    )
    commercial_warranty_length = first_non_empty_value(
        ws, headers, existing_variants, "Commercial Warranty Length"
    )

    return SheetContext(
        sheet_name=ws.title,
        headers=headers,
        existing_variants=existing_variants,
        product_name=product_name,
        base_code=base_code,
        image_links=image_links,
        color=color,
        commercial_warranty=commercial_warranty,
        commercial_warranty_length=commercial_warranty_length,
        dropbox_image_missing=dropbox_image_missing,
        dropbox_image_error=dropbox_image_error,
    )


def find_next_blank_row(ws: Worksheet, headers: dict[str, int]) -> int:
    """Find the next writable row based on Supplier Part Number."""

    part_column = headers["Supplier Part Number"]
    for row in range(DATA_START_ROW, ws.max_row + 1):
        value = ws.cell(row, part_column).value
        if value in (None, ""):
            return row
    return ws.max_row + 1


def format_size_label(width: int, height: int) -> str:
    """Format a Wayfair size label for the variant grouping."""

    return f'{height}" L x {width}" W'


def make_part_number(base_code: str, width: int, height: int, suffix: str) -> str:
    """Build a supplier/manufacturer part number for a new material variant."""

    return f"{base_code} {width}x{height} {suffix}"


def write_variant_groupings(
    ws: Worksheet,
    headers: dict[str, int],
    row: int,
    size_label: str,
    material_display_name: str,
) -> None:
    """Write Size and Wallpaper Material grouping values."""

    grouping_columns = [
        (
            f"Variant Grouping {index}",
            f"Variant Attribute Name On Site {index}",
        )
        for index in range(1, 4)
    ]
    existing_groupings = [
        get_cell_value(ws, headers, row, grouping_header)
        for grouping_header, value_header in grouping_columns
    ]

    if any(
        is_variant_grouping(grouping, "Size")
        or is_variant_grouping(grouping, "Wallpaper Material")
        for grouping in existing_groupings
    ):
        for grouping_header, value_header in grouping_columns:
            grouping = get_cell_value(ws, headers, row, grouping_header)
            if is_variant_grouping(grouping, "Size"):
                set_cell_value(ws, headers, row, grouping_header, "Size")
                set_cell_value(ws, headers, row, value_header, size_label)
            elif is_variant_grouping(grouping, "Wallpaper Material"):
                set_cell_value(ws, headers, row, grouping_header, "Wallpaper Material")
                set_cell_value(ws, headers, row, value_header, material_display_name)
        return

    set_cell_value(ws, headers, row, "Variant Grouping 1", "Size")
    set_cell_value(ws, headers, row, "Variant Attribute Name On Site 1", size_label)
    set_cell_value(ws, headers, row, "Variant Grouping 2", "Wallpaper Material")
    set_cell_value(
        ws,
        headers,
        row,
        "Variant Attribute Name On Site 2",
        material_display_name,
    )


def write_generated_row(
    ws: Worksheet,
    row: int,
    context: SheetContext,
    material_name: str,
    width: int,
    height: int,
    price_provider: PriceProvider,
    shaper: WallpaperDataShaper,
    include_technical_images: bool,
) -> list[str]:
    """Write all supported data cells for one generated variant row."""

    headers = context.headers
    attributes = shaper.print_type[material_name]
    price = price_provider.get_wallpaper_prices(width, height)[material_name]
    package = shaper.set_size_and_weight(height, width)
    primary_images = normalize_image_urls(shaper.clean_image_links(context.image_links))
    technical_images = (
        normalize_image_urls(shaper.technical_images)
        if include_technical_images
        else []
    )
    image_slots, additional_images = shaper.resolve_image_slots(
        primary_images,
        technical_images,
    )
    part_number = make_part_number(
        context.base_code,
        width,
        height,
        attributes.part_number_suffix,
    )

    values: dict[str, Any] = {
        "Supplier Part Number": part_number,
        "Manufacturer Part Number": part_number,
        "Product Name": context.product_name,
        "Base Cost": price,
        "Ship Type": "Small Parcel",
        "Freight Class": 400,
        "Lead Time": 120,
        "Replacement Lead Time": 120,
        "Carton Weight": package.weight,
        "Carton Height": package.height,
        "Carton Width": package.width,
        "Carton Depth": package.depth,
        "Warning Required": "No",
        "Country Of Manufacturer": "United States",
        "Image File Name or URL 1": image_slots[0],
        "Image File Name or URL 2": image_slots[1],
        "Image File Name or URL 3": image_slots[2],
        "Image File Name or URL 4": image_slots[3],
        "Image File Name or URL 5": image_slots[4],
        "Product Type": "Wall Mural",
        "Pattern": "Does Not Apply",
        "Wallpaper Texture": "Smooth",
        "Application Type": attributes.application,
        "Match Type": "Random",
        "Removal Type": attributes.removal,
        "Supplier Intended and Approved Use": "Non Residential Use; Residential Use",
        "BPA Free": "No",
        "Durability": (
            "Mold / Mildew Resistant; Water Resistant; Fade Resistant; "
            "Heat Resistant; Non-Porous; Non-Staining"
        ),
        "Designer": "Does Not Apply",
        "Product Care": "Wipe clean with a damp cloth",
        "Color": context.color,
        "Pattern Repeat Frequency": 0,
        "Wood Species": "Does Not Apply",
        "Sports Team Name": "Does Not Apply",
        "Movie / Show Series Name": "Does Not Apply",
        "Pattern Interval": 0.0,
        "Material": attributes.material,
        "Age Group": "All Ages",
        "Finish": "Primed",
        "Overall Product Length - End to End": shaper.convert_inches_to_feet(height),
        "Overall Width - Side to Side": width,
        "Square Footage per Unit": shaper.calculate_sq_ft(width, height),
        "Overall Product Weight": package.weight,
        "Commercial Warranty": context.commercial_warranty or "Yes",
        "Commercial Warranty Length": context.commercial_warranty_length or "30 Days",
        "Wayfair Compliance Verified Program (including Baby Safety Alliance fka JPMA) for this product category": "No",
        "Uniform Packaging and Labeling Regulations (UPLR) Compliant": "Yes",
        "Canada Product Restriction": "No",
        "Reason for Restriction": "Does Not Apply",
        "Sustainability & Social Responsibility Certifications (North America Only)": "No",
    }

    for header, value in values.items():
        set_cell_value(ws, headers, row, header, value)

    write_variant_groupings(
        ws,
        headers,
        row,
        format_size_label(width, height),
        material_grouping_label(material_name, shaper),
    )
    return additional_images


def write_additional_images(
    ws: Worksheet,
    part_number: str,
    image_urls: list[str],
) -> None:
    """Append additional image rows for overflow technical images."""

    if not image_urls:
        return
    headers = make_headers(ws)
    row = find_next_blank_row(ws, headers)
    for image_url in image_urls:
        set_cell_value(ws, headers, row, "Supplier Part Number", part_number)
        set_cell_value(ws, headers, row, "Image File Name or URL", image_url)
        row += 1


def fill_sheet(
    ws: Worksheet,
    context: SheetContext,
    materials: list[str],
    price_provider: PriceProvider,
    shaper: WallpaperDataShaper,
    additional_images_sheet: Worksheet | None,
    fill_existing: bool,
    write_overflow_images: bool,
    include_technical_images: bool,
) -> tuple[int, list[FillWarning]]:
    """Fill one SKU sheet and return added row count with warnings."""

    warnings: list[FillWarning] = []
    existing_part_numbers = {
        variant.part_number for variant in context.existing_variants
    }
    sizes = sorted(
        {(variant.width, variant.height) for variant in context.existing_variants}
    )
    next_row = find_next_blank_row(ws, context.headers)
    added_rows = 0

    if not context.image_links:
        warnings.append(
            FillWarning(
                sheet_name=context.sheet_name,
                message="No main image was found; Image File Name or URL 1 is blank.",
            )
        )
    elif context.dropbox_image_missing:
        warnings.append(
            FillWarning(
                sheet_name=context.sheet_name,
                message=(
                    f"No Dropbox image matched '{context.base_code}'. "
                    "Used another image source instead."
                ),
            )
        )

    if context.dropbox_image_missing and not context.image_links:
        if context.dropbox_image_error:
            warnings.append(
                FillWarning(
                    sheet_name=context.sheet_name,
                    message=(
                        f"Dropbox image matched '{context.base_code}', "
                        f"but link could not be resolved: {context.dropbox_image_error}"
                    ),
                )
            )
        warnings.append(
            FillWarning(
                sheet_name=context.sheet_name,
                message=f"No usable Dropbox image found for '{context.base_code}'.",
            )
        )

    if fill_existing:
        fill_existing_rows(
            ws,
            context,
            price_provider,
            shaper,
            include_technical_images,
        )

    for width, height in sizes:
        for material_name in materials:
            attributes = shaper.print_type[material_name]
            part_number = make_part_number(
                context.base_code,
                width,
                height,
                attributes.part_number_suffix,
            )
            if part_number in existing_part_numbers:
                continue
            if next_row > ws.max_row:
                warnings.append(
                    FillWarning(
                        sheet_name=context.sheet_name,
                        message="No free rows left in this SKU sheet.",
                    )
                )
                return added_rows, warnings

            additional_images = write_generated_row(
                ws=ws,
                row=next_row,
                context=context,
                material_name=material_name,
                width=width,
                height=height,
                price_provider=price_provider,
                shaper=shaper,
                include_technical_images=include_technical_images,
            )
            if write_overflow_images and additional_images_sheet is not None:
                write_additional_images(
                    additional_images_sheet,
                    part_number,
                    additional_images,
                )

            existing_part_numbers.add(part_number)
            next_row += 1
            added_rows += 1

    return added_rows, warnings


def fill_existing_rows(
    ws: Worksheet,
    context: SheetContext,
    price_provider: PriceProvider,
    shaper: WallpaperDataShaper,
    include_technical_images: bool,
) -> None:
    """Optionally fill supported cells in already existing rows."""

    for variant in context.existing_variants:
        if variant.material_name not in shaper.print_type:
            continue
        write_generated_row(
            ws=ws,
            row=variant.row_number,
            context=context,
            material_name=variant.material_name,
            width=variant.width,
            height=variant.height,
            price_provider=price_provider,
            shaper=shaper,
            include_technical_images=include_technical_images,
        )


def validate_materials(materials: list[str], shaper: WallpaperDataShaper) -> None:
    """Reject unknown material names before editing the workbook."""

    unknown = [material for material in materials if material not in shaper.print_type]
    if unknown:
        known = ", ".join(shaper.print_type)
        unknown_text = ", ".join(unknown)
        raise ValueError(f"Unknown material(s): {unknown_text}. Known: {known}.")


def fill_workbook(
    template_path: Path,
    output_path: Path,
    materials: list[str],
    image_map: dict[str, list[str]],
    dropbox_resolver: DropboxImageResolver | None,
    price_provider: PriceProvider,
    default_color: str,
    fill_existing: bool,
    write_overflow_images: bool,
    include_technical_images: bool,
) -> None:
    """Fill a Wayfair workbook and save it to a new file."""

    shaper = WallpaperDataShaper()
    validate_materials(materials, shaper)
    workbook = load_workbook(template_path)
    additional_images_sheet = (
        workbook["Additional Images"]
        if "Additional Images" in workbook.sheetnames
        else None
    )

    total_added = 0
    all_warnings: list[FillWarning] = []
    for ws in workbook.worksheets:
        if not is_sku_sheet(ws):
            continue
        context = make_sheet_context(ws, image_map, dropbox_resolver, default_color)
        if context is None:
            all_warnings.append(
                FillWarning(sheet_name=ws.title, message="No usable existing variants.")
            )
            continue
        added_rows, warnings = fill_sheet(
            ws=ws,
            context=context,
            materials=materials,
            price_provider=price_provider,
            shaper=shaper,
            additional_images_sheet=additional_images_sheet,
            fill_existing=fill_existing,
            write_overflow_images=write_overflow_images,
            include_technical_images=include_technical_images,
        )
        total_added += added_rows
        all_warnings.extend(warnings)
        print(f"{ws.title}: added {added_rows} variants")

    workbook.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Total added variants: {total_added}")
    if all_warnings:
        print("Warnings:")
        for warning in all_warnings:
            print(f"- {warning.sheet_name}: {warning.message}")


def main() -> None:
    """Run the command-line interface."""

    parser = build_parser()
    args = parser.parse_args()
    template_path = args.template.expanduser().resolve()
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else default_output_path(template_path)
    )
    price_provider = (
        PriceProvider(sheet_csv_url=args.price_csv_url)
        if args.price_csv_url
        else PriceProvider()
    )
    try:
        dropbox_resolver = make_dropbox_resolver(
            folder_path=args.dropbox_folder,
            token=args.dropbox_token,
            app_key=args.dropbox_app_key,
            app_secret=args.dropbox_app_secret,
            refresh_token=args.dropbox_refresh_token,
            recursive=not args.dropbox_no_recursive,
        )
    except (RuntimeError, ValueError) as exc:
        parser.error(str(exc))

    if args.dropbox_list_files:
        if dropbox_resolver is None:
            parser.error("--dropbox-list-files requires --dropbox-folder.")
        dropbox_resolver.print_visible_files()
        return

    fill_workbook(
        template_path=template_path,
        output_path=output_path,
        materials=args.materials,
        image_map=load_image_map(args.image_map),
        dropbox_resolver=dropbox_resolver,
        price_provider=price_provider,
        default_color=args.default_color,
        fill_existing=args.fill_existing,
        write_overflow_images=args.write_additional_images,
        include_technical_images=not args.skip_technical_images,
    )


if __name__ == "__main__":
    main()
