import datetime
import os

from openpyxl import load_workbook


class ExcelWriter:
    def __init__(
        self,
        template_filename: str,
        sheet_name: str,
    ) -> None:
        self.template_filename = template_filename
        self.sheet_name = sheet_name
        self.start_row: int = 6

    def write_data(self, new_data: list[dict], sku: str, folder: os.PathLike) -> None:
        wb = load_workbook(self.template_filename)
        ws = wb[self.sheet_name]
        headers = [cell.value for cell in ws[3] if cell.value]

        for i, row_data in enumerate(new_data, start=self.start_row):
            for col_idx, header in enumerate(headers, start=1):
                if header in row_data:
                    ws.cell(row=i, column=col_idx, value=row_data[header])

        filename = f"{sku}_{datetime.datetime.now().strftime('%d_%m_%Y_%H_%M')}.xlsx"
        if not os.path.exists(folder):
            os.makedirs(folder)
        full_path = os.path.join(folder, filename)
        wb.save(full_path)
