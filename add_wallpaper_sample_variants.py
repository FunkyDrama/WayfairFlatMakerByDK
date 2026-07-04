"""Fill Wayfair templates with 8x10 wallpaper sample variants only."""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from add_wallpaper_variants import (
    DEFAULT_DROPBOX_FOLDER,
    DEFAULT_MULTICOLOR_VALUE,
    FillWarning,
    find_next_blank_row,
    is_sku_sheet,
    load_image_map,
    make_dropbox_resolver,
    make_part_number,
    make_sheet_context,
    safe_filename_part,
    validate_materials,
    write_additional_images,
    write_generated_row,
)
from data.pricing import PriceProvider
from data.wallpaper_shaper import WallpaperDataShaper

SAMPLE_WIDTH = 8
SAMPLE_HEIGHT = 10
SAMPLE_MATERIALS = (
    "Peel-n-Stick",
    "Non-Woven",
    "Peel-n-Stick: Canvas",
    "Non-Woven: Premium",
)
TEMPLATE_GLOB = "Product Addition Template*.xlsx"


def runtime_directory() -> Path:
    """Return the directory containing the executable or this script."""

    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def find_adjacent_template(directory: Path) -> Path:
    """Find the newest Wayfair template next to the executable/script."""

    templates = sorted(
        (
            path
            for path in directory.glob(TEMPLATE_GLOB)
            if path.is_file() and not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not templates:
        raise FileNotFoundError(
            f"No '{TEMPLATE_GLOB}' file found next to {directory}."
        )
    return templates[0]


def collect_sku_sheet_names(input_path: Path) -> list[str]:
    """Return SKU sheet names from a Wayfair workbook."""

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        return [
            worksheet.title
            for worksheet in workbook.worksheets
            if is_sku_sheet(worksheet)
        ]
    finally:
        workbook.close()


def default_output_path(input_path: Path) -> Path:
    """Return the default sample-variant output path."""

    sku_names = [
        safe_filename_part(name) for name in collect_sku_sheet_names(input_path)
    ]
    if sku_names:
        return input_path.with_name(
            f"{'_'.join(sku_names)}_8x10_samples{input_path.suffix}"
        )
    return input_path.with_name(f"{input_path.stem}_8x10_samples{input_path.suffix}")


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser."""

    parser = argparse.ArgumentParser(
        description=(
            "Fill the newest adjacent Wayfair Product Addition Template with "
            "8x10 wallpaper sample variants for every wallpaper material."
        )
    )
    parser.add_argument(
        "--template",
        type=Path,
        help=(
            "Optional template path. By default the newest "
            f"'{TEMPLATE_GLOB}' next to the executable is used."
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Optional output .xlsx path. Defaults to '<SKUs>_8x10_samples.xlsx'.",
    )
    parser.add_argument(
        "--image-map",
        type=Path,
        help="Optional CSV with sheet/base_sku/sku and image1..image5 columns.",
    )
    parser.add_argument(
        "--dropbox-folder",
        default=DEFAULT_DROPBOX_FOLDER,
        help="Dropbox folder path or shared folder URL with main images.",
    )
    parser.add_argument(
        "--dropbox-token",
        help="Short-lived Dropbox API access token.",
    )
    parser.add_argument(
        "--dropbox-app-key",
        help="Dropbox app key for refresh-token auth.",
    )
    parser.add_argument(
        "--dropbox-app-secret",
        help="Dropbox app secret for refresh-token auth.",
    )
    parser.add_argument(
        "--dropbox-refresh-token",
        help="Dropbox refresh token.",
    )
    parser.add_argument(
        "--dropbox-no-recursive",
        action="store_true",
        help="Do not search Dropbox subfolders.",
    )
    parser.add_argument(
        "--price-csv-url",
        default=None,
        help="Optional Google Sheets CSV export URL for prices.",
    )
    parser.add_argument(
        "--default-color",
        default=DEFAULT_MULTICOLOR_VALUE,
        help="Color value used when the template has no existing Color answer.",
    )
    parser.add_argument(
        "--write-additional-images",
        action="store_true",
        help="Write overflow technical images to the Additional Images sheet.",
    )
    parser.add_argument(
        "--include-technical-images",
        action="store_true",
        help="Include default wallpaper technical images after primary images.",
    )
    return parser


def fill_sample_sheet(
    ws: Worksheet,
    context,
    price_provider: PriceProvider,
    shaper: WallpaperDataShaper,
    additional_images_sheet: Worksheet | None,
    write_overflow_images: bool,
    include_technical_images: bool,
) -> tuple[int, list[FillWarning]]:
    """Add missing 8x10 sample variants for one SKU sheet."""

    warnings: list[FillWarning] = []
    existing_part_numbers = {
        variant.part_number for variant in context.existing_variants
    }
    next_row = find_next_blank_row(ws, context.headers)
    added_rows = 0

    if not context.image_links:
        warnings.append(
            FillWarning(
                sheet_name=context.sheet_name,
                message="No main image was found; Image File Name or URL 1 is blank.",
            )
        )
    elif context.dropbox_image_missing:
        warnings.append(
            FillWarning(
                sheet_name=context.sheet_name,
                message=(
                    f"No Dropbox image matched '{context.base_code}'. "
                    "Used another image source instead."
                ),
            )
        )

    if context.dropbox_image_missing and not context.image_links:
        if context.dropbox_image_error:
            warnings.append(
                FillWarning(
                    sheet_name=context.sheet_name,
                    message=(
                        f"Dropbox image matched '{context.base_code}', "
                        f"but link could not be resolved: {context.dropbox_image_error}"
                    ),
                )
            )
        warnings.append(
            FillWarning(
                sheet_name=context.sheet_name,
                message=f"No usable Dropbox image found for '{context.base_code}'.",
            )
        )

    for material_name in SAMPLE_MATERIALS:
        attributes = shaper.print_type[material_name]
        part_number = make_part_number(
            context.base_code,
            SAMPLE_WIDTH,
            SAMPLE_HEIGHT,
            attributes.part_number_suffix,
        )
        if part_number in existing_part_numbers:
            continue
        if next_row > ws.max_row:
            warnings.append(
                FillWarning(
                    sheet_name=context.sheet_name,
                    message="No free rows left in this SKU sheet.",
                )
            )
            return added_rows, warnings

        additional_images = write_generated_row(
            ws=ws,
            row=next_row,
            context=context,
            material_name=material_name,
            width=SAMPLE_WIDTH,
            height=SAMPLE_HEIGHT,
            price_provider=price_provider,
            shaper=shaper,
            include_technical_images=include_technical_images,
        )
        if write_overflow_images and additional_images_sheet is not None:
            write_additional_images(
                additional_images_sheet,
                part_number,
                additional_images,
            )

        existing_part_numbers.add(part_number)
        next_row += 1
        added_rows += 1

    return added_rows, warnings


def fill_workbook(
    template_path: Path,
    output_path: Path,
    image_map: dict[str, list[str]],
    dropbox_resolver,
    price_provider: PriceProvider,
    default_color: str,
    write_overflow_images: bool,
    include_technical_images: bool,
) -> None:
    """Fill a Wayfair workbook with 8x10 sample variants and save it."""

    shaper = WallpaperDataShaper()
    validate_materials(list(SAMPLE_MATERIALS), shaper)
    workbook = load_workbook(template_path)
    additional_images_sheet = (
        workbook["Additional Images"]
        if "Additional Images" in workbook.sheetnames
        else None
    )

    total_added = 0
    all_warnings: list[FillWarning] = []
    for ws in workbook.worksheets:
        if not is_sku_sheet(ws):
            continue
        context = make_sheet_context(ws, image_map, dropbox_resolver, default_color)
        if context is None:
            all_warnings.append(
                FillWarning(sheet_name=ws.title, message="No usable existing variants.")
            )
            continue

        added_rows, warnings = fill_sample_sheet(
            ws=ws,
            context=context,
            price_provider=price_provider,
            shaper=shaper,
            additional_images_sheet=additional_images_sheet,
            write_overflow_images=write_overflow_images,
            include_technical_images=include_technical_images,
        )
        total_added += added_rows
        all_warnings.extend(warnings)
        print(f"{ws.title}: added {added_rows} sample variants")

    workbook.save(output_path)
    print(f"Template: {template_path}")
    print(f"Saved: {output_path}")
    print(f"Total added sample variants: {total_added}")
    if all_warnings:
        print("Warnings:")
        for warning in all_warnings:
            print(f"- {warning.sheet_name}: {warning.message}")


def main() -> None:
    """Run the command-line interface."""

    parser = build_parser()
    args = parser.parse_args()
    base_dir = runtime_directory()
    template_path = (
        args.template.expanduser().resolve()
        if args.template
        else find_adjacent_template(base_dir)
    )
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else default_output_path(template_path)
    )
    price_provider = (
        PriceProvider(sheet_csv_url=args.price_csv_url)
        if args.price_csv_url
        else PriceProvider()
    )

    try:
        dropbox_resolver = make_dropbox_resolver(
            folder_path=args.dropbox_folder,
            token=args.dropbox_token,
            app_key=args.dropbox_app_key,
            app_secret=args.dropbox_app_secret,
            refresh_token=args.dropbox_refresh_token,
            recursive=not args.dropbox_no_recursive,
        )
        fill_workbook(
            template_path=template_path,
            output_path=output_path,
            image_map=load_image_map(args.image_map),
            dropbox_resolver=dropbox_resolver,
            price_provider=price_provider,
            default_color=args.default_color,
            write_overflow_images=args.write_additional_images,
            include_technical_images=args.include_technical_images,
        )
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        parser.error(str(exc))


if __name__ == "__main__":
    main()
