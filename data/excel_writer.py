"""Workbook writing helpers for Wayfair templates."""

import datetime
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
RowData = dict[str, Any]


class ExcelWriter:
    """Write generated data into the Excel template."""

    def __init__(self, sheet_name: str) -> None:
        """Store template metadata for subsequent writes."""

        self.template_filename = Path(ASSETS_DIR / "template.xlsx")
        self.sheet_name = sheet_name
        self.start_row: int = 8

    @staticmethod
    def write_sheet(
        ws: Worksheet,
        start_row: int,
        header_row: int,
        data: list[RowData],
    ) -> None:
        """Write rows into a worksheet using header names as keys."""

        headers = [cell.value for cell in ws[header_row] if cell.value]

        for i, row_data in enumerate(data, start=start_row):
            for col_idx, header in enumerate(headers, start=1):
                if header in row_data:
                    ws.cell(row=i, column=col_idx, value=row_data[header])

    @staticmethod
    def build_output_path(sku: str, folder: str | os.PathLike[str]) -> Path:
        """Build and create the target path for a generated workbook."""

        filename = f"{sku}_{datetime.datetime.now().strftime('%d_%m_%Y_%H_%M')}.xlsx"
        folder_path = Path(folder)
        folder_path.mkdir(parents=True, exist_ok=True)
        return folder_path / filename

    def write_data(
        self,
        new_data: list[RowData],
        sku: str,
        folder: str | os.PathLike[str],
    ) -> None:
        """Write the main sheet into a new workbook file."""

        wb = load_workbook(self.template_filename)
        ws = wb[self.sheet_name]
        self.write_sheet(ws, self.start_row, 4, new_data)
        wb.save(self.build_output_path(sku, folder))

    def write_data_with_additional_images(
        self,
        new_data: list[RowData],
        additional_images_data: list[RowData],
        sku: str,
        folder: str | os.PathLike[str],
    ) -> None:
        """Write the main sheet and optional additional-images sheet."""

        wb = load_workbook(self.template_filename)
        self.write_sheet(wb[self.sheet_name], self.start_row, 4, new_data)

        if additional_images_data:
            self.write_sheet(
                wb["Additional Images"],
                start_row=7,
                header_row=4,
                data=additional_images_data,
            )
        wb.save(self.build_output_path(sku, folder))
